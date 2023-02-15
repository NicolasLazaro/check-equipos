import os
import os.path
import shutil
from openpyxl import workbook
from openpyxl import load_workbook

src_folder = input('Path to check: ')
dest_folder = r'C:\Users\nmlazaro\Documents\project\sigue-checklist'
nocheck = r'C:\Users\nmlazaro\Documents\project\no-sigue-checklist'


sheet_mant = 'Mantenimiento Gral'
sheet_equipos = 'Equipos'

wb = load_workbook('Equipos.xlsx', data_only=True) #Abro el archivo excel que voy a completar
ws = wb['Equipos']

#El acc global revisa la ultima linea escrita del archivo y suma 1 para seguir desde alli

acc = ws.max_row + 1

def tableIsEmpty(sheet, columns):
    empty = True
    for i in range(8,18):
        if rowIsEmpty(sheet, i, columns):
            empty = True
            continue
        else:
            empty = False
            break
    return empty

#Verifica que la fila tenga contenido no repetido, si se repite es porque esta vacio.
def rowIsEmpty(sheet, rows, columns):
    list = []
    for i in columns:
        list.append(str(sheet[str(i)+str(rows)].value))
    result = all(element == list[0] for element in list)
    if result:
        return True
    else:
        return False


#Verifica si existe la hoja que pasemos por parametro, si existe debe extraer la info y pegarla en el Excel de informacion
def app(sheet):
    global acc
     #Acumulador para no pisar las rows una vez que finaliza un archivo
    for file in os.listdir(src_folder):
        if file.endswith('xlsx') or file.endswith('xls'):
            try:
                loaded_wb = load_workbook(os.path.join(src_folder, file), data_only=True)
                loaded_ws = loaded_wb['Equipos']
                if 'Equipos' in loaded_wb.sheetnames:
                    print('este es el archivo '+ str(file))
                    #Extraigo info de la sheet 'Equipos' de todos los archivos en la raiz y la llevo a mi otro archivo
                    data_columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
                    if not tableIsEmpty(loaded_ws, data_columns):
                        for rows in range(8, 50):
                            
                            if not rowIsEmpty(loaded_ws, rows, data_columns):
                                for columns in data_columns:
                                    ws[columns+str(acc)].value = loaded_ws[columns+str(rows)].value
                                    #Escribo en la celda [L,acc] el nombre del archivo con su ruta respectiva
                                    ws['L'+(str(acc))].value = str(os.path.join(src_folder, file))

                                
                                acc += 1
                    else:
                        ws['B'+(str(acc))].value = 'Sin equipos'
                        ws['L'+(str(acc))].value = str(os.path.join(src_folder, file))
                        acc += 1
                    
                    for i in range(2, ws.max_row):
                        if rowIsEmpty(ws, i, data_columns):
                            ws.delete_rows(i, 1)
                            #print('DEL '+str(i)+' in '+str(ws))
                        # else:
                        #     print(str(i) + ' Sin tocar')

                    wb.save('Equipos.xlsx')
            #Handle para no cortar el programa cuando el key 'Equipos' no se encuentre
            except KeyError:
                continue

        else:
            shutil.copy2(os.path.join(src_folder, file), nocheck) #Si no cumple se copian en una carpeta
            print('Algunos archivos no cumplen con lo solicitado, se copiaron a la carpeta no-sigue-checklist')
            continue

answer = True

while(answer):
    app(sheet_equipos)

    ask_value = int(input('Desea seguir con otra ruta?\n[1]Si\n[2]No\n'))
    if (ask_value == 2):
        answer = False
        print('Proceso terminado.')
    elif (ask_value == 1):
        src_folder = input('Path to check: ')
    else:
        print('Valor incorrecto como respuesta.')
        break




